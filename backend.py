"""Functions for writing data to a sqlite database, an Outlook email and an Excel sheet"""

import sqlite3

import win32com.client as win32

from dynaconf import Dynaconf
from openpyxl import load_workbook


def _dict_to_html_table(data):
    """Utility to convert a dictionary to an HTML table, for neater display in an email"""

    fields = "".join([f"<td>{k}</td>" for k in data.keys()])
    values = "".join([f"<td>{v}</td>" for v in data.values()])

    return f"<table border=1><tr>{fields}</tr><tr>{values}</tr><table>"


def _write_to_db(config, data):
    """Write the given data to a sqlite database"""

    with sqlite3.connect(config.db.path) as conn:
        fields = ",".join(data.keys())
        values = ",".join([f"'{v}'" for v in data.values()])
        cur = conn.cursor()
        cur.execute(f"INSERT INTO {config.db.table}({fields}) VALUES ({values})")
        conn.commit()


def _write_to_excel(config, data):
    """Append the given data to the first empty row of an Excel sheet"""

    wb = load_workbook(filename=config.excel.path)
    ws = wb[config.excel.sheet]
    first_empty_row_no = ws.max_row + 1

    for col_no, val in enumerate(data.values()):
        ws.cell(column=col_no + 1, row=first_empty_row_no, value=val)

    wb.save(filename=config.excel.path)


def _write_to_outlook(config, data):
    """Launch MS Outlook and populate a draft email containing the given data"""

    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = config.outlook.recipient
    mail.Subject = config.outlook.subject
    mail.HtmlBody = "Added the following data:\n" + _dict_to_html_table(data)
    mail.display(False)


def write_data(data):
    """
        Write data (captured as a dictionary of fields:values to each of sqlite, Outlook and Excel
        N.B. config.yaml contains the various configurable parameters for the app
        (filepaths, email address, etc.)
    """

    config = Dynaconf(settings_files=["config.yaml"])
    _write_to_db(config, data)
    _write_to_excel(config, data)
    _write_to_outlook(config, data)
